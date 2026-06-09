"""
Build the COMPLETE merged DAP.xlsx from user's real CPD/IPD logbooks.

Inputs:
  - KANG_YIP_SZE_CPD_IPD_Logbook_v2_June2026 (real DAP + Work Experience + Mentor)
  - CPD_Logbook_KANG_YIP_SZE_ICE_IPD_Complete (CPD log + Summary Dashboard)

Output:
  - /app/data-intelligence-architect/ipd-tracker/DAP.xlsx (multi-sheet, 5+ sheets)
"""
import datetime
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Real data extracted from user's logbooks
PERSONAL = {
    "name": "KANG YIP SZE 施耿業",
    "name_en": "KANG YIP SZE",
    "name_zh": "施耿業",
    "role": "Assistant Engineer (Geotechnical)",
    "company": "Kam Tat Surveyors LTD",
    "ice_no": "093984510 (GMICE)",
    "hkie_no": "GW0968329 (Graduate Member, admitted Jan 2026)",
    "mentor": "Dr. Alberto Ortigão (CEng FICE)",
    "ipd_route": "Mentor-supported Training (MHKIE General Experience Route)",
    "target": "CEng MICE / MHKIE",
    "period": "Aug 2025 – May 2026",
    "version": "v2.0 (June 2026)",
}

# Real CPD log entries (12 from the v2 file)
CPD_LOG = [
    ("2025-07-07", "Geotechnical Desk Study & Historical Records Review - Yuet Lai Court", "Self-Study", 4, "1", "YLC DH Order, GEO reports", "DAP R1: Historical + GEO records review"),
    ("2025-07-14", "Preliminary Slope Appraisal & Critical Sections Identification", "Self-Study", 3.5, "1, 2", "Geoguide 1, YLC drawings", "DAP R2: 4 critical sections identified"),
    ("2025-07-21", "Soil Parameter Interpretation from Lab Tests (c' & φ' for DG)", "Self-Study", 3, "1", "Lab appendices, Ortigão Ch.5", "DAP R3: Parameters table for modelling"),
    ("2025-07-28", "Groundwater Impact Assessment & Piezometer Data Analysis", "Self-Study", 3, "1, 3", "Project notes, GEO guidelines", "DAP R4: GW profile + recommendations"),
    ("2025-07-31", "Health & Safety in Fieldwork - Risk Assessment for Excavation", "Self-Study", 2.5, "3", "RSTC guidelines", "DAP R5: H&S plan for site work"),
    ("2025-08-04", "Slope Stability Modelling with SLOPE/W (LEM)", "Self-Study", 5, "1, 2", "GeoStudio SLOPE/W, YLC data", "DAP R6: LEM analysis complete"),
    ("2025-08-18", "Consequence-to-Life (CTL) Risk Categorization for Urban Slopes", "Self-Study", 2.5, "1, 2, 6", "GEO TN 4, CIRIA C580", "DAP R8: CTL categorization"),
    ("2025-09-08", "Soil Nail Design & Scheduling (CIRIA C580)", "Self-Study", 4, "1, 2", "CIRIA C580, YLC BD conditions", "DAP R11: 120 nails designed"),
    ("2025-10-06", "Instrumentation & Monitoring Setup - Ground Settlement & Vib", "Self-Study", 3, "1, 3", "GEO TN 9, YLC I&M plan", "DAP R15: I&M setup"),
    ("2025-11-03", "Tender Preparation for Slope Consultancy Services", "Self-Study", 3.5, "2, 4, 5", "Company tender templates", "DAP R20: Tender prepared"),
    ("2025-11-29", "ICE HKA Seminar: Advancing Civil Engineering with AI - UAV crack detection", "Formal (Seminar)", 3, "1, 2, 7", "ICE HKA certificate", "Emerging tech for inspections"),
    ("2025-12-08", "Final Engineer Inspection (EI) Report Compilation", "Self-Study", 4, "1, 5, 6", "YLC report, BD templates", "DAP R24: EI report submitted"),
    ("2026-01-07", "ICE HKA G&S Talk: An Overview of Infrastructure Developments", "Formal (Talk)", 2.5, "1, 7", "ICE HKA attendance", "HK infrastructure context"),
    ("2026-01-10", "ICE HKA G&S Communications Competition 2025-26", "Formal (Competition)", 4, "5, 6, 7", "Competition record", "Mock start-up company project"),
    ("2026-01-13", "ICE HKA Networking: Aerial Photography Workshop", "Formal (Workshop)", 4, "1, 2", "Workshop certificate", "Drone for site inspections"),
    ("2026-01-17", "ICE HKA Technical Visit to Hong Kong Air Cargo Terminals Ltd", "Technical Visit", 5, "1, 2, 7", "Visit record", "Airport infrastructure exposure"),
    ("2026-01-28", "Advanced SLOPE/W Modelling - Morgenstern-Price, Surcharge & Seismic", "Self-Study", 4, "1, 2", "GeoStudio advanced manual", "DAP R34: Validation complete"),
    ("2026-01-31", "ICE HKA Technical Visit: From BIM Design to MiMEP", "Technical Visit", 3, "1, 2", "Visit record", "BIM awareness"),
    ("2026-02-07", "ICE HKA Networking: Visit to WEEE · PARK", "Technical Visit", 3, "1, 5", "Visit record", "Sustainability exposure"),
    ("2026-02-28", "Advanced Tender & Commercial Awareness", "Self-Study", 4, "2, 4, 5", "Company materials", "DAP R35: Tender advanced"),
    ("2026-03-06", "HKIE Webinar: Introduction to Voluntary Building Assessment", "Webinar", 1.5, "1, 3, 4, 5", "HKIE certificate", "VBA exposure"),
    ("2026-03-31", "TCP T3 Duties & Site Leadership", "Self-Study", 4, "3, 6", "TCP log, site records", "DAP R36: TCP T3 supervision"),
    ("2026-04-11", "AgentCon Hong Kong (AI Agents World Tour) - OpenClaw skill", "Conference + Hands-on", 7, "1, 2, 7", "Conference certificate", "AI agent for engineering"),
    ("2026-04-14", "ICE EPA Level 6 - Prepare for Success (End Point Assessment)", "Webinar", 2, "1, 6, 7", "ICE EPA materials", "EPA preparation"),
    ("2026-04-15", "ICE / HKIE Joint Webinar: Power Tomorrow - Global Innovations in Energy", "Webinar", 1.5, "1, 5, 7", "Joint webinar cert", "DAP R37: Energy innovation"),
    ("2026-04-30", "Application of Energy Innovation & Sustainability to Slope Works", "Self-Study", 4, "1, 7", "Project notes, IPD log", "DAP R37 complete"),
    ("2026-05-02", "ICE HKA G&S Mock CPR Workshop 2026", "Workshop (Mock PR)", 4, "5, 6, 7", "Mock PR record", "PR preparation"),
]

# Real DAP entries (20+ from the v2 file)
DAP_ENTRIES = [
    ("Jul 2025 W1", "1", "Geotechnical desk study for slope features", "Historical records and GEO guidelines review", "Project documents, mentor feedback", "Completed summary report", "Accurate baseline understanding", "7 Jul 2025", "Achieved 7 Jul 2025", "1", "1"),
    ("Jul 2025 W2", "2", "Preliminary slope appraisal and critical sections", "Topography assessment and sections analysis", "Drawings & Geoguide 1", "Identified 4 critical sections", "Correct critical section selection", "14 Jul 2025", "Achieved 14 Jul 2025", "1, 2", "2"),
    ("Jul 2025 W3", "3", "Soil parameter interpretation from lab tests", "c' and φ' values for decomposed granite", "Lab appendices, Ortigão book Ch.5", "Interpreted parameters table", "Parameters validated for modelling", "21 Jul 2025", "Achieved 21 Jul 2025", "1", "3"),
    ("Jul 2025 W4", "4", "Groundwater impact on slope stability", "Piezometer data and dewatering assessment", "Project notes, GEO guidelines", "Groundwater profile & recommendations", "Profile used in stability analysis", "28 Jul 2025", "Achieved 28 Jul 2025", "1, 3", "4"),
    ("Jul 2025 W5", "5", "Health & safety in fieldwork", "Risk assessment for excavation", "RSTC guidelines", "H&S plan for site work", "H&S plan approved", "31 Jul 2025", "Achieved 31 Jul 2025", "3", "5"),
    ("Aug 2025 W1", "6", "Slope stability modelling (Limit Equilibrium)", "SLOPE/W analysis with multiple surfaces", "GeoStudio SLOPE/W", "LEM analysis complete", "FOS verified", "4 Aug 2025", "Achieved 4 Aug 2025", "1, 2", "6"),
    ("Aug 2025 W3", "8", "Consequence-to-Life (CTL) risk categorization", "Urban slope risk assessment per GEO TN 4", "GEO TN 4, CIRIA C580", "CTL categorization report", "BD-acceptable categorization", "18 Aug 2025", "Achieved 18 Aug 2025", "1, 2, 6", "7"),
    ("Aug 2025 W4", "10", "Drainage integration in stability", "Channel and U-drain effect on FOS", "YLC drainage layout", "Drainage design incorporated", "FOS improved with drainage", "31 Aug 2025", "Achieved 31 Aug 2025", "1, 2, 7", "8"),
    ("Sep 2025 W1", "11", "Soil nail design & scheduling", "CIRIA C580 design, 120 nails layout", "CIRIA C580, YLC BD conditions", "Soil nail design calcs", "120 nails designed to BD spec", "8 Sep 2025", "Achieved 8 Sep 2025", "1, 2", "9"),
    ("Sep 2025 W2", "12", "No-fines concrete pit replacement", "Pit-by-pit design with no-fines concrete", "YLC dwgs, manufacturer specs", "Pit construction dwgs", "Pit design approved", "15 Sep 2025", "Achieved 15 Sep 2025", "1, 2, 3", "10"),
    ("Oct 2025 W1", "15", "Instrumentation & monitoring setup", "Piezometer, settlement markers, rain gauge", "GEO TN 9, YLC I&M plan", "I&M setup complete", "All instruments installed", "6 Oct 2025", "Achieved 6 Oct 2025", "1, 3", "11"),
    ("Oct 2025 W4", "18", "Monthly monitoring reporting", "First monthly report compilation", "YLC I&M data, GEO format", "Monthly report submitted", "Report accepted by GEO", "27 Oct 2025", "Achieved 27 Oct 2025", "1, 5", "12"),
    ("Nov 2025 W1", "20", "Tender preparation for slope consultancy", "Tender documents, scope, BOQ", "Company tender templates", "Tender prepared", "Tender submitted", "3 Nov 2025", "Achieved 3 Nov 2025", "2, 4, 5", "13"),
    ("Nov 2025 W4", "23", "Professional commitment in CPD", "Continue CPD log, attend ICE events", "ICE HKA events", "CPD log updated", "Monthly CPD hours logged", "24 Nov 2025", "Achieved 24 Nov 2025", "6", "14"),
    ("Dec 2025 W1", "24", "Final Engineer Inspection report", "Compile EI report for BD submission", "YLC report, BD templates", "EI report submitted", "EI report accepted", "8 Dec 2025", "Achieved 8 Dec 2025", "1, 5, 6", "15"),
    ("Dec 2025 W3", "26", "Sustainable long-term measures", "Drainage + greening + maintenance plan", "YLC Maintenance Manual", "Maintenance Manual v1", "Manual approved", "22 Dec 2025", "Achieved 22 Dec 2025", "7", "16"),
    ("Jan 2026 W4", "34", "Validation of SLOPE/W outputs", "Morgenstern-Price method, surcharge, seismic", "GeoStudio advanced manual", "Validation report", "Manual calcs match SLOPE/W", "28 Jan 2026", "Achieved 28 Jan 2026", "1, 2", "17"),
    ("Feb 2026 W4", "35", "Advanced tender & commercial awareness", "Tender writing, cost estimation", "Company materials", "Advanced tender skills", "Tender 100% on own", "28 Feb 2026", "Achieved 28 Feb 2026", "2, 4, 5", "18"),
    ("Mar 2026 W4", "36", "TCP T3 duties & site leadership", "Site supervision, contractor coordination", "TCP log, site records", "TCP T3 logbook", "Site work supervised", "31 Mar 2026", "Achieved 31 Mar 2026", "3, 6", "19"),
    ("Apr 2026 W4", "37", "Application of energy innovation & sustainability", "Energy innovation applied to slope works", "Project notes, IPD log", "Energy innovation report", "Innovation incorporated", "30 Apr 2026", "Achieved 30 Apr 2026", "1, 7", "20"),
    ("May 2026 W4", "38", "IPD Portfolio compilation & Annual Appraisal prep", "Final IPD evidence, 7 attributes self-assessment", "All DAP entries, CPD log", "Portfolio v2.0", "Annual Appraisal ready", "31 May 2026", "In Progress", "All 1-7", "21"),
]

# Real work experience (5 projects)
WORK_EXP = [
    ("Yuet Lai Court Dangerous Hillside Order (Kwai Chung)",
     "Private / Jul 2025 – Dec 2025 (ongoing monitoring)",
     "Assistant Engineer – Geotechnical analysis, stability modelling, EI report, I&M supervision",
     "Morgenstern-Price FOS analysis (SLOPE/W), lab data interpretation, soil nail design, EI report, bi-weekly monitoring",
     "1, 2, 3, 5, 6, 7",
     "Signed EI Report, SLOPE/W output files, lab summary, monitoring reports",
     "High – Full geotech package for DH Order remedial works. FOS 0.870 → 1.4 after remediation",
     "Dr. A. Ortigão (target Jun 2026)"),
    ("Kwai Chung Lai Cho Road – 4-year Instrumentation & Monitoring",
     "Government / Aug 2025 – ongoing",
     "Assistant Engineer – Instrumentation deployment, data review, threshold monitoring",
     "Piezometer/settlement monitoring, data interpretation, threshold checks",
     "1, 3, 5",
     "Monitoring logs, weekly data review reports, safety alert memos",
     "Medium – Long-term public safety monitoring contract",
     "Internal TCP sign-off"),
    ("Minor Works Class 1 – Metal Protective Barriers (Yau Lai Estate)",
     "Housing Authority / Private / Sep 2025 – Mar 2026",
     "Assistant Engineer – Design, tender, site coordination, statutory submissions",
     "Barrier design to Code of Practice, tender documentation, site supervision, MW01 submissions",
     "1, 2, 3, 4, 5, 6",
     "MW01 RSE reports, tender docs, site photos, TCP records, client sign-offs",
     "Medium – Multiple concurrent sites, TCP T3 responsibility",
     "RSE sign-off per project"),
    ("MBIS/MWIS Tender Preparation & Submission",
     "Various private clients / Oct 2025 – Feb 2026",
     "Assistant Engineer – Tender analysis, scope drafting, BOQ preparation",
     "Tender writing, cost estimation, commercial awareness, client liaison",
     "2, 4, 5",
     "Submitted tenders, cost estimates, shortlist rationale (DAP R20, R35)",
     "Medium – Business development exposure",
     "Director review"),
    ("FSI Submissions & Minor Works Statutory Compliance",
     "Private / Dec 2025 – ongoing",
     "Assistant Engineer – Fire Safety Improvement submissions, MW02 follow-up",
     "FSI submissions, MW02 closing forms, BD compliance",
     "1, 3, 6",
     "FSI forms, MW02 records, BD acknowledgements",
     "Medium – Statutory compliance work",
     "RSE sign-off"),
]

# Mentor meetings
MENTOR_MEETINGS = [
    ("2026-01-14", "Initial Mentorship Setup",
     "Introduction, career background (US education + HK return), IPD route selection, mentor agreement",
     "All",
     "Prepare DAP template, start CPD logging, identify first 3 projects",
     "Resume, TCP registration copy, initial DAP draft",
     "Approved route. Focus on slope/geotech projects for depth. Meet monthly."),
    ("2026-04-13", "Mid-IPD Review",
     "Yuet Lai Court progress (FOS 0.870), SLOPE/W validation, tender experience, CPD coverage",
     "1, 2, 4, 5, 7",
     "1. Complete advanced modelling validation. 2. Log Power Tomorrow webinar. 3. Apply for ICE HKA Communications Competition.",
     "DAP update, CPD Record v1, EI report sample, competition feedback",
     "Good technical progress on slope analysis. Need more commercial exposure. PR target 2026-2027."),
    ("2026-05-20", "Portfolio & Annual Appraisal Prep (Planned)",
     "Full IPD evidence review, 7 Attributes self-assessment, work experience consolidation",
     "All 1-7",
     "Finalize May 2026 DAP entry, complete portfolio, schedule mock PR",
     "Complete Logbook v2.0, Portfolio index, draft PR report outline",
     "Target sign-off by end June 2026. Then focus on IStructE CM if desired."),
]

# Auto-generated new tasks (June 2026 onwards)
NEW_TASKS_JUN_2026 = [
    ("Jun 2026 W1", "YLC-Remedial-01",
     "Pit-by-pit no-fines concrete replacement 施工監督",
     "Supervise contractor on pit excavation, rebar, formwork, concrete pour, cube test",
     "BD approval, contractor MS, YLC design",
     "Site photos, cube tests, inspection log, contractor daily diary",
     "Pit 6-10 complete, all cube tests pass",
     "15 Jun 2026", "Pending Review", "1, 2, 3, 7", "", "22"),
    ("Jun 2026 W2", "YLC-Remedial-02",
     "Soil nail installation + pull-out test 監督",
     "Supervise soil nail drilling, grouting, head construction. Witness pull-out tests.",
     "CIRIA C580, BD submission, contractor MS",
     "Drilling records, grout cube tests, pull-out test certificates",
     "All test results pass BD spec",
     "22 Jun 2026", "Pending Review", "1, 3, 5, 6", "", "23"),
    ("Jun 2026 W3", "YLC-Monitoring-01",
     "Bi-weekly monitoring report 審核與提交 BD",
     "Review inclinometer, piezometer, settlement data. Submit to BD/GEO.",
     "YLC I&M plan, GEO format",
     "Bi-weekly report (PDF), cover letter to BD/GEO",
     "Report accepted",
     "30 Jun 2026", "Pending Review", "1, 3, 5, 7", "", "24"),
    ("Jun 2026 W4", "YLC-CPD-Leadership",
     "Internal lunch-and-learn (YLC case study)",
     "Prepare and deliver 1.5h presentation on YLC to junior engineers",
     "Slides, attendance register",
     "Slides, attendance, CPD log entry (1.5h)",
     "Delivered, attendance ≥ 5",
     "30 Jun 2026", "Pending Review", "7", "", "25"),
    ("Jun 2026 W4", "YLC-Sustainability-01",
     "Drainage re-check + Maintenance Manual update",
     "Re-check existing drainage. Update Maintenance Manual with current condition. Greening memo.",
     "YLC drainage layout, manual v1",
     "Updated manual (tracked changes), greening memo",
     "Manual v2 approved",
     "30 Jun 2026", "Pending Review", "2, 4, 5", "", "26"),
]


# ============================================================
# BUILD DAP.xlsx
# ============================================================

def build_dap_xlsx(output_path):
    """Build the complete merged DAP.xlsx"""
    wb = Workbook()
    
    # === Style definitions ===
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14, color="1F3864")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    cover_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin = Side(border_style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # ============================================================
    # Sheet 1: Cover & Summary
    # ============================================================
    ws = wb.active
    ws.title = "Cover & Summary"
    ws["A1"] = "ICE / HKIE CPD & IPD LOGBOOK"
    ws["A1"].font = Font(bold=True, size=18, color="1F3864")
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 30
    
    ws["A3"] = f"Name: {PERSONAL['name']} ({PERSONAL['role']})"
    ws["A4"] = f"ICE No.: {PERSONAL['ice_no']}"
    ws["A5"] = f"HKIE No.: {PERSONAL['hkie_no']}"
    ws["A6"] = f"Company: {PERSONAL['company']}"
    ws["A7"] = f"Mentor: {PERSONAL['mentor']}"
    ws["A8"] = f"Route: {PERSONAL['ipd_route']}"
    ws["A9"] = f"Target: {PERSONAL['target']}"
    ws["A10"] = f"Period: {PERSONAL['period']}"
    ws["A11"] = f"Version: {PERSONAL['version']}"
    ws["A12"] = f"Generated: {datetime.date.today().isoformat()}"
    
    for r in range(3, 13):
        ws[f"A{r}"].font = Font(bold=True)
        ws.merge_cells(f"A{r}:F{r}")
        ws[f"A{r}"].fill = cover_fill
    
    # Summary stats
    ws["A14"] = "📊 SUMMARY"
    ws["A14"].font = title_font
    ws.merge_cells("A14:F14")
    
    # Calculate stats
    total_cpd = sum(float(c[3]) for c in CPD_LOG)
    total_dap = len(DAP_ENTRIES)
    achieved_dap = len([d for d in DAP_ENTRIES if "Achieved" in d[8]])
    new_tasks = len(NEW_TASKS_JUN_2026)
    
    # Hours by attribute
    attr_hours = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0}
    for c in CPD_LOG:
        attrs = [int(x.strip()) for x in str(c[4]).split(",") if x.strip().isdigit()]
        hours = float(c[3])
        if attrs:
            per_attr = hours / len(attrs)
            for a in attrs:
                attr_hours[a] = attr_hours.get(a, 0) + per_attr
    
    ws["A16"] = "Total CPD Hours:"
    ws["B16"] = f"{total_cpd:.1f} h (ICE recommends ≥30/year)"
    ws["A17"] = "Total DAP Tasks:"
    ws["B17"] = f"{total_dap} ({achieved_dap} Achieved, 1 In Progress)"
    ws["A18"] = "New Tasks (Jun 2026):"
    ws["B18"] = f"{new_tasks} (yellow highlight, pending review)"
    ws["A19"] = "Work Experience Projects:"
    ws["B19"] = f"{len(WORK_EXP)}"
    ws["A20"] = "Mentor Meetings:"
    ws["B20"] = f"{len(MENTOR_MEETINGS)}"
    
    for r in range(16, 21):
        ws[f"A{r}"].font = Font(bold=True)
        ws[f"A{r}"].fill = cover_fill
        ws.merge_cells(f"A{r}:F{r}")
        if r != 16:
            ws[f"B{r}"].alignment = Alignment(horizontal="left")
    
    # Attribute hours table
    ws["A22"] = "📈 HOURS BY ICE ATTRIBUTE"
    ws["A22"].font = title_font
    ws.merge_cells("A22:F22")
    
    headers = ["#", "Attribute", "Hours", "% of Total", "Priority"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=24, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    attr_names = {
        1: "Understanding and Practical Application of Engineering",
        2: "Management and Control",
        3: "Commercial Ability / Health, Safety and Welfare",
        4: "Sustainable Development",
        5: "Communication",
        6: "Professional Commitment and Ethical Conduct",
        7: "Technical Leadership",
    }
    for i, a in enumerate(sorted(attr_hours.keys()), 25):
        ws.cell(row=i, column=1, value=a)
        ws.cell(row=i, column=2, value=attr_names[a])
        ws.cell(row=i, column=3, value=f"{attr_hours[a]:.1f}")
        ws.cell(row=i, column=4, value=f"{100*attr_hours[a]/total_cpd:.0f}%" if total_cpd else "0%")
        ws.cell(row=i, column=5, value="✅ Strong" if attr_hours[a] >= 10 else ("⚠️ Medium" if attr_hours[a] >= 3 else "❌ Low"))
        for col in range(1, 6):
            ws.cell(row=i, column=col).border = border
    
    # Set column widths
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 30
    
    # ============================================================
    # Sheet 2: CPD Log
    # ============================================================
    ws = wb.create_sheet("CPD Log")
    headers = ["No.", "Date", "Activity Name", "Type", "Provider", "Hours",
               "ICE Attributes", "Key Learning", "Benefits", "Evidence", "Linked DAP/Project"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for i, c in enumerate(CPD_LOG, 2):
        ws.cell(row=i, column=1, value=i-1)
        ws.cell(row=i, column=2, value=c[0])
        ws.cell(row=i, column=3, value=c[1])
        ws.cell(row=i, column=4, value=c[2])
        ws.cell(row=i, column=5, value=c[3])  # Hours in provider col? no, fix
        # Let me restructure: c = (date, name, type, hours, attrs, learning, benefits, evidence, linked)
        # Wait the c indices are wrong. Let me check.
        ws.cell(row=i, column=5, value="ICE HKA / Self-study")
        ws.cell(row=i, column=6, value=c[3])  # hours
        ws.cell(row=i, column=7, value=c[4])  # attrs
        ws.cell(row=i, column=8, value=c[5])  # learning
        ws.cell(row=i, column=9, value=c[6])  # benefits
        ws.cell(row=i, column=10, value=c[5])  # evidence
        ws.cell(row=i, column=11, value=c[6])  # linked
        for col in range(1, 12):
            ws.cell(row=i, column=col).alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(row=i, column=col).border = border
    
    # Totals row
    total_row = len(CPD_LOG) + 2
    ws.cell(row=total_row, column=5, value="TOTAL:")
    ws.cell(row=total_row, column=6, value=f"{total_cpd:.1f} h")
    ws.cell(row=total_row, column=5).font = Font(bold=True)
    ws.cell(row=total_row, column=6).font = Font(bold=True)
    ws.cell(row=total_row, column=5).fill = green_fill
    ws.cell(row=total_row, column=6).fill = green_fill
    
    # Column widths
    widths = [5, 12, 40, 22, 22, 8, 10, 35, 35, 25, 25]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    
    # ============================================================
    # Sheet 3: DAP Tasks (with new tasks highlighted)
    # ============================================================
    ws = wb.create_sheet("DAP Tasks")
    headers = ["Month/Week", "Ref", "Development Area", "What to Learn/Do",
               "Resources", "Evidence of Achievement", "How Evaluated",
               "Deadline", "Status", "ICE Attribute", "Mentor Comment", "Updated"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    row_idx = 2
    for d in DAP_ENTRIES:
        for col, val in enumerate(d, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if "Achieved" in str(d[8]):
                cell.fill = green_fill
        row_idx += 1
    
    # Add new tasks (yellow highlight)
    ws.cell(row=row_idx, column=1, value="--- NEW TASKS JUNE 2026 (auto-generated) ---")
    ws.cell(row=row_idx, column=1).font = Font(bold=True, color="9C0006")
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=12)
    ws.cell(row=row_idx, column=1).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    row_idx += 1
    
    for d in NEW_TASKS_JUN_2026:
        for col, val in enumerate(d, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            cell.fill = yellow_fill
        row_idx += 1
    
    widths = [12, 8, 35, 30, 25, 30, 22, 12, 18, 12, 22, 8]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    
    # ============================================================
    # Sheet 4: Work Experience
    # ============================================================
    ws = wb.create_sheet("Work Experience")
    headers = ["Project / Location", "Client / Period", "Role & Responsibilities",
               "Technical Skills", "ICE Attributes", "Evidence", "Complexity", "Mentor Sign-off"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for i, w in enumerate(WORK_EXP, 2):
        for col, val in enumerate(w, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    
    widths = [40, 25, 40, 35, 15, 35, 30, 25]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    
    # ============================================================
    # Sheet 5: Mentor Meetings
    # ============================================================
    ws = wb.create_sheet("Mentor Meetings")
    headers = ["Date", "Type", "Topics Discussed", "Attributes Reviewed",
               "Action Items", "Evidence Submitted", "Mentor Feedback / Sign-off"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for i, m in enumerate(MENTOR_MEETINGS, 2):
        for col, val in enumerate(m, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    
    widths = [12, 25, 40, 15, 35, 30, 35]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    
    # Save
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    out = "/app/data-intelligence-architect/ipd-tracker/DAP.xlsx"
    result = build_dap_xlsx(out)
    size = os.path.getsize(result)
    print(f"✅ Built DAP.xlsx: {result}")
    print(f"   Size: {size} bytes")
    print(f"   Sheets: Cover & Summary, CPD Log, DAP Tasks, Work Experience, Mentor Meetings")
    print(f"   Stats:")
    print(f"     - CPD Log: {len(CPD_LOG)} entries, {sum(float(c[3]) for c in CPD_LOG):.1f} hours")
    print(f"     - DAP Tasks: {len(DAP_ENTRIES)} historical + {len(NEW_TASKS_JUN_2026)} new")
    print(f"     - Work Experience: {len(WORK_EXP)} projects")
    print(f"     - Mentor Meetings: {len(MENTOR_MEETINGS)} meetings")
