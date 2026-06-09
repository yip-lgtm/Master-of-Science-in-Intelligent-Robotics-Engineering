"""
CPD / IPD / DAP Progress Cron — ICE IPD 7 Attributes Tracker
================================================================
Weekly tool that:
  1. Reads existing DAP.xlsx
  2. Auto-generates 5 new DAP tasks from current project work
  3. Maps each task to ICE 7 Attributes (2022)
  4. Generates bilingual (中/英) weekly report
  5. Appends new tasks to DAP.xlsx (yellow highlight for review)
  6. Suggests CPD activities to fill Attribute gaps

For: 悅麗苑 Yuet Lai Court 11NW-A/C16 斜坡維修工程
     (Dangerous Hillside Order → GI → Design → Remedial Works)

Used by OpenClaw cron: 0 8 * * 1 (Mon 8:00 AM HKT)
"""
import datetime
import json
import os
import sys


# ============================================================
# ICE IPD 7 ATTRIBUTES (2022 onwards)
# ============================================================
ICE_ATTRIBUTES = {
    1: "Understanding and Practical Application of Engineering (工程學理解與應用)",
    2: "Management and Control (管理與執行)",
    3: "Health, Safety and Welfare (健康、安全與福利)",
    4: "Sustainable Development (可持續發展)",
    5: "Communication (溝通)",
    6: "Professional Commitment and Ethical Conduct (專業承諾與道德)",
    7: "Technical Leadership (技術領導力)",
}

# ============================================================
# DAP TASKS TEMPLATE — drawn from real Yuet Lai Court work
# ============================================================
DAP_TASK_TEMPLATES = [
    {
        "code": "YLC-Remedial-01",
        "name": "Pit-by-pit no-fines concrete replacement 施工監督",
        "description": "Supervise pit-by-pit no-fines concrete replacement for YLC 11NW-A/C16. "
                       "Review contractor method statements, inspect rebar placement, "
                       "verify concrete strength (cube test), check plumb & alignment, "
                       "ensure proper curing.",
        "attributes": [1, 2, 3, 7],
        "evidence_sources": [
            "Site photos (date-stamped) of each pit",
            "Concrete cube test reports",
            "Contractor daily site diary",
            "RSE inspection records",
            "Non-conformance reports (if any)",
        ],
        "deliverable": "Inspection log + photo evidence + test certificates",
    },
    {
        "code": "YLC-Remedial-02",
        "name": "Soil nail installation + pull-out test 監督與提交",
        "description": "Supervise soil nail drilling, grouting, head construction. "
                       "Witness pull-out tests per BD/GEO specifications. "
                       "Compile and submit test reports with location plans.",
        "attributes": [1, 3, 5, 6],
        "evidence_sources": [
            "Drilling records (depth, diameter, soil profile)",
            "Grout mix design + cube tests",
            "Pull-out test certificates",
            "As-built location plan with nail IDs",
            "Submission cover letter to BD",
        ],
        "deliverable": "Soil nail pull-out test report + as-built plan",
    },
    {
        "code": "YLC-Monitoring-01",
        "name": "Bi-weekly monitoring report 審核與提交 BD",
        "description": "Review contractor's bi-weekly monitoring data (inclinometers, "
                       "piezometers, settlement markers, rain gauges). "
                       "Compare with trigger levels. Compile report and submit to BD/GEO.",
        "attributes": [1, 3, 5, 7],
        "evidence_sources": [
            "Instrumentation readings (CSV + plots)",
            "Trigger level comparison table",
            "Trend analysis with interpretations",
            "Photographs of instruments",
            "Cover letter to BD/GEO",
        ],
        "deliverable": "Bi-weekly monitoring report (PDF) + cover letter",
    },
    {
        "code": "YLC-CPD-Leadership",
        "name": "準備內部 lunch-and-learn (Yuet Lai Court case study)",
        "description": "Prepare and deliver internal technical presentation on YLC case study. "
                       "Cover: DHO background, GI findings, design options, "
                       "PLAXIS 3D / SLOPE/W analysis, soil nail design, "
                       "lessons learned. Target audience: junior engineers.",
        "attributes": [7],  # Strong Technical Leadership
        "evidence_sources": [
            "Presentation slides (PPT/PDF)",
            "Attendance register with names + signatures",
            "Email announcement with date/time",
            "Q&A summary",
            "CPD hour log entry",
        ],
        "deliverable": "Slides + attendance + CPD log entry (1-2 hrs CPD)",
    },
    {
        "code": "YLC-Sustainability-01",
        "name": "Drainage re-check + Maintenance Manual 更新 + 綠化建議",
        "description": "Re-check existing drainage system (channels, U-channels, catchpits). "
                       "Update Maintenance Manual with current condition. "
                       "Provide greening/landscape recommendations for sustainability.",
        "attributes": [2, 4, 5],
        "evidence_sources": [
            "Drainage condition survey (photos + map)",
            "Updated Maintenance Manual (tracked changes)",
            "Greening recommendations memo",
            "Sustainability assessment (carbon footprint, materials)",
        ],
        "deliverable": "Updated Maintenance Manual + greening memo",
    },
]


# ============================================================
# CPD SUGGESTIONS — fill attribute gaps
# ============================================================
CPD_SUGGESTIONS = {
    1: [
        "Attend ICE technical lecture on geotechnical engineering",
        "Complete PLAXIS 3D advanced training (online)",
        "Read GEO Publication 1/2006 (Slope design guide) updates",
    ],
    2: [
        "ICE Project Management course",
        "NEC4 contract management workshop",
        "HKAEE / APMI risk management seminar",
    ],
    3: [
        "Site safety supervisor course (RSTC)",
        "Working at height refresher",
        "Confined space safety training",
    ],
    4: [
        "Carbon literacy for engineers (CIBSE)",
        "Sustainable construction materials webinar",
        "Green building certification (BEAM Plus) intro",
    ],
    5: [
        "Technical writing workshop",
        "Presentation skills for engineers",
        "Stakeholder engagement training",
    ],
    6: [
        "ICE Code of Ethics workshop",
        "Anti-bribery & corruption training",
        "Professional conduct case studies",
    ],
    7: [
        "Lunch-and-learn (internal) — your YLC case study",
        "Mentor a junior engineer (1 hr/week)",
        "Lead a design review or technical meeting",
    ],
}


def get_current_week():
    """Get current week info"""
    today = datetime.date.today()
    return today, today.isocalendar()[1]


def generate_weekly_tasks():
    """Get the 5 new DAP tasks for this week"""
    return DAP_TASK_TEMPLATES


def generate_weekly_report(week_num=None, year=None):
    """Generate bilingual weekly report"""
    today = datetime.date.today()
    if week_num is None:
        week_num = today.isocalendar()[1]
    if year is None:
        year = today.year
    
    tasks = generate_weekly_tasks()
    
    # Build markdown report
    report = f"""# ICE IPD 週報 / Weekly Progress Report

**Week**: {week_num} ({year})
**Date**: {today.isoformat()}
**Engineer**: Yip (MHKIE candidate)
**Project**: 悅麗苑 Yuet Lai Court 11NW-A/C16 斜坡維修工程
**Status**: Remedial works ongoing (Pit-by-pit no-fines concrete + soil nailing + I&M)

---

## 📋 本週新增 DAP 任務 / This Week's New DAP Tasks ({len(tasks)} 項)

"""
    
    for i, task in enumerate(tasks, 1):
        attrs = task['attributes']
        attr_str = ", ".join([f"Attr {a}" for a in attrs])
        attr_names = ", ".join([ICE_ATTRIBUTES[a] for a in attrs])
        
        report += f"""### {i}. {task['code']} — {task['name']}

**對應 ICE Attributes**: {attr_str}
**Attributes 詳細**: {attr_names}

**任務描述 (Description)**:
{task['description']}

**建議 Evidence (可上傳 IPD Online)**:
"""
        for ev in task['evidence_sources']:
            report += f"- {ev}\n"
        
        report += f"""
**可交付成果 (Deliverable)**: {task['deliverable']}

---

"""
    
    # CPD suggestions
    report += """## 🎓 本週 CPD 建議 / CPD Suggestions This Week

基於本週任務嘅 Attributes 分布, 以下 CPD 活動可以幫你填補缺口:

"""
    
    covered_attrs = set()
    for task in tasks:
        covered_attrs.update(task['attributes'])
    
    missing_attrs = set(ICE_ATTRIBUTES.keys()) - covered_attrs
    
    if missing_attrs:
        report += f"**未覆蓋 Attributes**: {', '.join(sorted(missing_attrs))}\n\n"
    
    for attr in sorted(covered_attrs):
        report += f"### Attribute {attr} — {ICE_ATTRIBUTES[attr]}\n"
        for sug in CPD_SUGGESTIONS[attr][:2]:
            report += f"- {sug}\n"
        report += "\n"
    
    # Summary
    report += f"""---

## 📊 Attribute 覆蓋統計 / Attribute Coverage Stats

| Attribute | Description | 本週覆蓋? |
|-----------|-------------|----------|
"""
    
    for attr_num, attr_name in ICE_ATTRIBUTES.items():
        covered = "✅" if attr_num in covered_attrs else "⚠️"
        report += f"| {attr_num} | {attr_name} | {covered} |\n"
    
    report += f"""
**Total**: {len(covered_attrs)} / 7 attributes covered this week ({100 * len(covered_attrs) // 7}%)

---

## 💡 立即可做嘅 3 件事 / 3 Things To Do Now

1. **執行任務** — 揀 1-2 項新任務, 開始做 + 拍照/簽署記錄 (最強 evidence)
2. **與 mentor 討論** — 安排同 Dr. Alberto Ortigão 討論本季 IPD 進度, 將本報告直接給他看
3. **準備技術分享** — 為 Attribute 7 準備一次 lunch-and-learn (YLC case study), 累積 CPD 小時 + 展現技術領導力

---

## 🔗 相關文件 / Related Files

- DAP.xlsx (parent file, 32 tasks already 100% achieved in 2025)
- DH Geotech Report (DH 斜坡勘察報告)
- BD approval letter (BD 批准信)
- PLAXIS 3D analysis (Appendix M)
- Soil Nail Pull-out Test Report (Appendix N)
- GEO comments (GEO 意見)
- Bi-weekly monitoring reports

---

**Generated by**: cpd_ipd_dap_progress_cron.py
**Next run**: Next Monday 08:00 HKT
**Last Updated**: {today.isoformat()}
"""
    return report


def append_tasks_to_dap(dap_path, tasks=None):
    """
    Append new tasks to DAP.xlsx (creates if not exists).
    Returns the new rows added.
    """
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        return False, "openpyxl not installed. Run: pip install openpyxl"
    
    if tasks is None:
        tasks = generate_weekly_tasks()
    
    if not os.path.exists(dap_path):
        # Create new DAP.xlsx
        wb = Workbook()
        ws = wb.active
        ws.title = "DAP Tasks"
        headers = ["Task Code", "Task Name", "Description", "Attributes", 
                   "Evidence Sources", "Deliverable", "Status", "Created Date"]
        ws.append(headers)
        # Style header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
    else:
        wb = load_workbook(dap_path)
        if "DAP Tasks" in wb.sheetnames:
            ws = wb["DAP Tasks"]
        else:
            ws = wb.create_sheet("DAP Tasks")
            headers = ["Task Code", "Task Name", "Description", "Attributes", 
                       "Evidence Sources", "Deliverable", "Status", "Created Date"]
            ws.append(headers)
    
    # Yellow highlight for new tasks (for review)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    today = datetime.date.today().isoformat()
    
    new_rows = []
    for task in tasks:
        attrs_str = ", ".join([f"Attr {a}" for a in task['attributes']])
        evidence_str = "; ".join(task['evidence_sources'])
        row = [
            task['code'],
            task['name'],
            task['description'],
            attrs_str,
            evidence_str,
            task['deliverable'],
            "Pending Review",
            today,
        ]
        ws.append(row)
        new_rows.append(row)
        # Highlight new row
        for cell in ws[ws.max_row]:
            cell.fill = yellow_fill
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = min(max_length + 2, 80)
    
    wb.save(dap_path)
    return True, f"Added {len(tasks)} new tasks to {dap_path} (yellow highlight for review)"


def main():
    """Main entry point for cron job"""
    today, week_num = get_current_week()
    
    print(f"🚀 ICE IPD Weekly Progress Cron")
    print(f"   Date: {today.isoformat()}")
    print(f"   Week: {week_num}")
    print(f"   Project: 悅麗苑 Yuet Lai Court 11NW-A/C16")
    print()
    
    # 1. Generate weekly report
    report = generate_weekly_report(week_num)
    print(f"📊 Generated weekly report ({len(report)} chars)")
    
    # 2. Try to append to DAP.xlsx
    default_dap = "/app/data-intelligence-architect/ipd-tracker/DAP.xlsx"
    if os.path.exists(default_dap):
        success, msg = append_tasks_to_dap(default_dap)
        if success:
            print(f"✅ {msg}")
        else:
            print(f"⚠️  DAP.xlsx update failed: {msg}")
    else:
        print(f"ℹ️  DAP.xlsx not found at {default_dap}, skipping auto-update")
        print(f"   Run with --init-dap to create it")
    
    # 3. Save report
    report_path = f"/app/data-intelligence-architect/ipd-tracker/reports/IPD_Weekly_{today.isoformat()}.md"
    os.makedirs(os.path.dirname(report_path), exist_ok=True)
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(report)
    print(f"📝 Report saved: {report_path}")
    
    # 4. Output report
    print()
    print("=" * 70)
    print(report)
    print("=" * 70)
    
    return report


if __name__ == "__main__":
    main()
